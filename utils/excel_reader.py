import openpyxl

class ExcelReader:
    global wb
    def __init__(self, file, sheet):
        wb = openpyxl.load_workbook(file)
        self.sheet = wb[sheet]

    def get_data(self):

        rows = []
        for r in range(2, self.sheet.max_row + 1):
            username = self.sheet.cell(r, 1).value
            password = self.sheet.cell(r, 2).value
            expected = self.sheet.cell(r, 3).value
            rows.append((username, password, expected))
        return rows

    def get_steps(self,testcase_no):
        steps = []
        self.sheet = wb[testcase_no]
        for r in range(2, self.sheet.max_row + 1):
            steps.append({
                "keyword": self.sheet.cell(r, 3).value,
                "page": self.sheet.cell(r, 4).value,
                "element": self.sheet.cell(r, 5).value,
                "data": self.sheet.cell(r, 6).value,
                "expected": self.sheet.cell(r, 7).value
            })
        return steps
