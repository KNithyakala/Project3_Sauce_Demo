import openpyxl

class KeywordReader:

    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file)

    def get_steps(self,testcase_no):
        steps = []
        sheet = self.wb[testcase_no]
        for r in range(2, sheet.max_row + 1):
            steps.append({
                "keyword": sheet.cell(r, 3).value,
                "page": sheet.cell(r, 4).value,
                "element": sheet.cell(r, 5).value,
                "data": sheet.cell(r, 6).value,
                "expected": sheet.cell(r, 7).value
            })
        return steps
